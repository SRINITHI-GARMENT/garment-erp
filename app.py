import os
import uuid
import json
from datetime import datetime
from collections import defaultdict
from functools import wraps
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, jsonify, session, send_file
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)

# ---------------- DB CONFIG ----------------
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "postgresql+psycopg2://postgres.cusbryojwnldabchhfkx:9788%40Srinithi@aws-1-ap-northeast-1.pooler.supabase.com:6543/postgres",
)
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True, "pool_recycle": 300}
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-change-me")

db = SQLAlchemy(app)


# ---------------- PERMISSIONS LIST ----------------
PERMISSION_GROUPS = [
    ("Colour Master", [
        ("colour_view","View"),("colour_add","Add"),
        ("colour_edit","Edit"),("colour_delete","Delete"),
    ]),
    ("Size Master", [
        ("size_view","View"),("size_add","Add"),
        ("size_edit","Edit"),("size_delete","Delete"),
    ]),
    ("Fabric Master", [
        ("fabric_view","View"),("fabric_add","Add"),
        ("fabric_edit","Edit"),("fabric_delete","Delete"),
    ]),
    ("Product Master", [
        ("product_view","View"),("product_add","Add"),
        ("product_edit","Edit"),("product_delete","Delete"),
    ]),
    ("Supplier Master", [
        ("supplier_view","View"),("supplier_add","Add"),
        ("supplier_edit","Edit"),("supplier_delete","Delete"),
    ]),
    ("Program Entry", [
        ("program_view","View"),("program_add","Add"),
        ("program_edit","Edit"),("program_delete","Delete"),
        ("program_status","Change Status / Proceed / Complete"),
    ]),
    ("Overall Programs", [
        ("overall_view","View"),("overall_status","Change Status"),
    ]),
    ("User Master", [
        ("user_view","View"),("user_add","Add"),
        ("user_edit","Edit"),("user_delete","Delete"),
    ]),
    ("Fabric Base Stock", [
        ("base_stock_view","View"),("base_stock_add","Add"),
        ("base_stock_edit","Edit"),("base_stock_delete","Delete"),
    ]),
    ("Fabric Actual Stock", [
        ("actual_stock_view","View"),("actual_stock_add","Add"),
        ("actual_stock_edit","Edit"),("actual_stock_delete","Delete"),
    ]),
    ("Fabric Requirement", [
        ("requirement_view","View"),("requirement_add","Add"),
        ("requirement_edit","Edit"),("requirement_delete","Delete"),
    ]),
    ("Fabric Orders", [
        ("fabric_orders_view","View"),("fabric_orders_add","Add"),
        ("fabric_orders_edit","Edit"),("fabric_orders_delete","Delete"),('fabric_orders_manual_add',"Manual order"),
    ]),
    ("Process Master", [
        ("process_view","View"),("process_add","Add"),
        ("process_edit","Edit"),("process_delete","Delete"),
    ]),
]
ALL_PERMISSIONS = [code for _, items in PERMISSION_GROUPS for code, _ in items]


# ---------------- MODELS ----------------
class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default="user", nullable=False)
    permissions_csv = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, raw): self.password_hash = generate_password_hash(raw)
    def check_password(self, raw): return check_password_hash(self.password_hash, raw)

    @property
    def permissions(self):
        return [p for p in (self.permissions_csv or "").split(",") if p]
    @permissions.setter
    def permissions(self, values):
        self.permissions_csv = ",".join(values or [])

    def has(self, perm):
        if self.role == "admin": return True
        return perm in self.permissions


class Colour(db.Model):
    __tablename__ = "colours"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(50), nullable=False)


class Size(db.Model):
    __tablename__ = "sizes"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)


class Fabric(db.Model):
    __tablename__ = "fabrics"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    uom = db.Column(db.String(50), nullable=False)
    gsm = db.Column(db.String(50), nullable=False)
    dia_csv = db.Column(db.Text, default="")
    colour_csv = db.Column(db.Text, default="")
    @property
    def dia(self): return [x for x in (self.dia_csv or "").split(",") if x]
    @dia.setter
    def dia(self, values): self.dia_csv = ",".join(values or [])
    @property
    def colour(self): return [x for x in (self.colour_csv or "").split(",") if x]
    @colour.setter
    def colour(self, values): self.colour_csv = ",".join(values or [])


class Product(db.Model):
    __tablename__ = "products"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    brand = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(120), nullable=False)
    type = db.Column(db.String(120), nullable=False)
    fabric = db.Column(db.String(120), nullable=False)
    colors_csv = db.Column(db.Text, default="")
    sizes_csv = db.Column(db.Text, default="")
    @property
    def colors(self): return [x for x in (self.colors_csv or "").split(",") if x]
    @colors.setter
    def colors(self, values): self.colors_csv = ",".join(values or [])
    @property
    def sizes(self): return [x for x in (self.sizes_csv or "").split(",") if x]
    @sizes.setter
    def sizes(self, values): self.sizes_csv = ",".join(values or [])


class Supplier(db.Model):
    __tablename__ = "suppliers"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(50), default="")
    contact = db.Column(db.String(120), default="")
    address = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Program(db.Model):
    __tablename__ = "programs"
    id = db.Column(db.String(64), primary_key=True)
    program_no = db.Column(db.String(20), nullable=False, index=True)
    date = db.Column(db.String(20), nullable=False)
    fabric = db.Column(db.String(120), nullable=False)
    dia = db.Column(db.String(50))
    type = db.Column(db.String(120))
    product = db.Column(db.String(120), nullable=False)
    color = db.Column(db.String(120), nullable=False)
    size = db.Column(db.String(50), nullable=False)
    ratio = db.Column(db.String(50), nullable=False)
    rolls = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(20), default="pending")  # pending / wip / completed
    supplier = db.Column(db.String(120), default="")
    lot_no = db.Column(db.String(50), default="")
    completed_date = db.Column(db.String(20), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class Counter(db.Model):
    __tablename__ = "counters"
    name = db.Column(db.String(50), primary_key=True)
    value = db.Column(db.Integer, default=0)


# -------- STOCK MANAGEMENT MODELS (from FAB NEW) --------
class StockEntry(db.Model):
    __tablename__ = 'stock_entry'
    id = db.Column(db.Integer, primary_key=True)
    fabric_id = db.Column(db.Integer, db.ForeignKey('fabrics.id'), nullable=False)
    entry_type = db.Column(db.String(50), nullable=False)  # base_stock, actual_stock, requirement
    quantity = db.Column(db.Float, nullable=False)
    note = db.Column(db.Text)
    colour_csv = db.Column('colour', db.Text, default='')  # CSV format stored in the DB column named colour
    dia_csv = db.Column('dia', db.Text, default='')  # CSV format stored in the DB column named dia
    min_purchase_qty = db.Column(db.Float)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    gsm = db.Column(db.String(50))
    
    fabric = db.relationship('Fabric', backref='stock_entries')
    
    @property
    def colour(self): return [x.strip() for x in (self.colour_csv or '').split(',') if x.strip()]
    @colour.setter
    def colour(self, values): self.colour_csv = ','.join(values or [])
    
    @property
    def dia(self): return [x.strip() for x in (self.dia_csv or '').split(',') if x.strip()]
    @dia.setter
    def dia(self, values): self.dia_csv = ','.join(values or [])


class GeneratedOrder(db.Model):
    __tablename__ = 'generated_orders'
    id = db.Column(db.Integer, primary_key=True)
    po_number = db.Column(db.String(255), nullable=False, unique=True)
    order_items = db.Column(db.Text, nullable=False)  # JSON array stored as text
    process_type = db.Column(db.String(255))
    process = db.Column(db.String(255))
    fabric_incharge = db.Column(db.String(255))
    status = db.Column(db.String(50), default='Pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class FabricIncharge(db.Model):
    __tablename__ = 'fabric_incharge'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False, unique=True)


class Process(db.Model):
    __tablename__ = 'process'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False, unique=True)


class ProcessDetails(db.Model):
    __tablename__ = 'process_details'
    id = db.Column(db.Integer, primary_key=True)
    process_type = db.Column(db.String(255), nullable=False)
    process_id = db.Column(db.Integer, db.ForeignKey('process.id'), nullable=False)
    fabric_incharge_id = db.Column(db.Integer, db.ForeignKey('fabric_incharge.id'), nullable=False)
    
    process = db.relationship('Process', backref='details')
    fabric_incharge = db.relationship('FabricIncharge', backref='process_details')


# -------- STOCK MANAGEMENT CONSTANTS --------
STOCK_TYPES = {
    'base_stock': 'Fabric Base Stock',
    'actual_stock': 'Fabric Actual Stock',
    'requirement': 'Requirement',
}

ENTRY_LABELS = {
    'base_stock': 'Base Stock',
    'actual_stock': 'Actual Stock',
    'requirement': 'Requirement',
}

ALLOWED_EXCEL_EXTENSIONS = {'xlsx'}


# ---------------- AUTH HELPERS ----------------
def current_user():
    uid = session.get("user_id")
    if not uid: return None
    return User.query.get(uid)

def login_required(view):
    @wraps(view)
    def wrapped(*a, **k):
        if not current_user(): return redirect("/")
        return view(*a, **k)
    return wrapped

def role_required(*roles):
    def deco(view):
        @wraps(view)
        def wrapped(*a, **k):
            u = current_user()
            if not u: return redirect("/")
            if u.role not in roles:
                return "Access denied.", 403
            return view(*a, **k)
        return wrapped
    return deco

def permission_required(perm):
    def deco(view):
        @wraps(view)
        def wrapped(*a, **k):
            u = current_user()
            if not u: return redirect("/")
            if not u.has(perm):
                return f"Access denied: missing permission '{perm}'.", 403
            return view(*a, **k)
        return wrapped
    return deco

@app.context_processor
def inject_user():
    return {"current_user": current_user()}


# ---------------- DICT HELPERS ----------------
def _to_dict_colour(c): return {"id":c.id,"name":c.name,"code":c.code}
def _to_dict_size(s):   return {"id":s.id,"name":s.name}
def _to_dict_fabric(f): return {"id":f.id,"name":f.name,"uom":f.uom,"gsm":f.gsm,"dia":f.dia,"colour":f.colour}
def _to_dict_product(p):
    return {"id":p.id,"name":p.name,"brand":p.brand,"category":p.category,
            "type":p.type,"fabric":p.fabric,"colors":p.colors,"sizes":p.sizes}
def _to_dict_supplier(s):
    return {"id":s.id,"name":s.name,"code":s.code,"contact":s.contact,
            "address":s.address,"created_at":s.created_at}

def _next_program_no():
    result = db.session.execute(text("SELECT nextval('program_no_seq')"))
    next_num = result.scalar()
    return f"PRG{next_num:03d}"


def _grouped_programs():
    rows = (Program.query
            .order_by(Program.created_at.desc(), Program.program_no, Program.color)
            .all())
    groups = {}; order = []
    for r in rows:
        key = (r.program_no, r.color)
        if key not in groups:
            groups[key] = {
                "ids": [], "program_no": r.program_no, "date": r.date,
                "fabric": r.fabric, "dia": r.dia, "type": r.type,
                "product": r.product, "color": r.color,
                "sizes": [], "ratios": [], "rolls": r.rolls, "statuses": [],
                "supplier": r.supplier or "", "lot_no": r.lot_no or "",
                "completed_date": r.completed_date or "",
            }
            order.append(key)
        groups[key]["ids"].append(r.id)
        groups[key]["sizes"].append(str(r.size))
        groups[key]["ratios"].append(str(r.ratio))
        groups[key]["statuses"].append((r.status or "pending").lower())

    out = []
    for key in order:
        g = groups[key]
        sts = g["statuses"]
        if sts and all(s == "completed" for s in sts):
            status = "completed"
        elif any(s == "wip" for s in sts):
            status = "wip"
        elif any(s == "completed" for s in sts):
            status = "wip"   # mixed → still in progress
        else:
            status = "pending"
        out.append({
            "ids": ",".join(g["ids"]),
            "id": g["ids"][0],
            "program_no": g["program_no"],
            "date": g["date"], "fabric": g["fabric"], "dia": g["dia"],
            "type": g["type"], "product": g["product"], "color": g["color"],
            "size": ":".join(g["sizes"]),
            "ratio": ":".join(g["ratios"]),
            "rolls": g["rolls"], "status": status,
            "supplier": g["supplier"], "lot_no": g["lot_no"],
            "completed_date": g["completed_date"],
        })
    return out


# -------- STOCK MANAGEMENT HELPERS --------
from openpyxl import Workbook, load_workbook

def allowed_excel_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS

def parse_list_field(field_name):
    values = [value.strip() for value in request.form.getlist(field_name) if value.strip()]
    return values

def request_list_arg(name):
    values = []
    for raw_value in request.args.getlist(name):
        values.extend(part.strip() for part in raw_value.split(',') if part.strip())
    return values

def request_filter_values(keys):
    return {key: request_list_arg(key) for key in keys}

def filter_values(selected_filters, key):
    value = (selected_filters or {}).get(key, [])
    if isinstance(value, str):
        return [part.strip() for part in value.split(',') if part.strip()]
    return [str(part).strip() for part in value if str(part).strip()]

def matches_filter(value, selected_values):
    selected_values = [str(v).strip().lower() for v in selected_values if str(v).strip()]
    if not selected_values:
        return True
    value_str = str(value or '').lower()
    return any(selected_value in value_str for selected_value in selected_values)

def matches_list_filter(values, selected_values):
    selected_values = [str(v).strip().lower() for v in selected_values if str(v).strip()]
    if not selected_values:
        return True
    for value in values or []:
        value_str = str(value or '').lower()
        if any(selected_value in value_str for selected_value in selected_values):
            return True
    return False

def build_filter_options(rows, include_result=False):
    options = {}
    options['fabric'] = sorted({row.get('fabric_name') for row in rows if row.get('fabric_name')})
    options['uom'] = sorted({row.get('uom') for row in rows if row.get('uom')})
    options['gsm'] = sorted({row.get('gsm') for row in rows if row.get('gsm')})
    options['colour'] = sorted({value for row in rows for value in row.get('colour', []) if value})
    options['dia'] = sorted({value for row in rows for value in row.get('dia', []) if value})
    if include_result:
        options['result'] = sorted({row.get('result_type') for row in rows if row.get('result_type')})
    return options

def filter_stock_rows(rows, selected_filters):
    filtered = []
    for row in rows:
        if not matches_filter(row.get('fabric_name'), filter_values(selected_filters, 'fabric')):
            continue
        if not matches_filter(row.get('uom'), filter_values(selected_filters, 'uom')):
            continue
        if not matches_filter(row.get('gsm'), filter_values(selected_filters, 'gsm')):
            continue
        if not matches_list_filter(row.get('colour'), filter_values(selected_filters, 'colour')):
            continue
        if not matches_list_filter(row.get('dia'), filter_values(selected_filters, 'dia')):
            continue
        filtered.append(row)
    return filtered


def get_fabrics():
    fabrics = Fabric.query.order_by(Fabric.name).all()
    return [_to_dict_fabric(f) for f in fabrics]


def get_stock_entries(entry_type, selected_filters=None):
    entries = StockEntry.query.filter(StockEntry.entry_type == entry_type).order_by(StockEntry.created_at.desc()).all()
    parsed_rows = []
    for entry in entries:
        fabric = entry.fabric
        row_dict = {
            'id': entry.id,
            'fabric_id': entry.fabric_id,
            'entry_type': entry.entry_type,
            'quantity': entry.quantity,
            'note': entry.note,
            'colour': entry.colour,
            'dia': entry.dia,
            'min_purchase_qty': entry.min_purchase_qty,
            'created_at': entry.created_at.isoformat() if entry.created_at else None,
            'fabric_name': fabric.name if fabric else None,
            'uom': fabric.uom if fabric else None,
            'gsm': entry.gsm or (fabric.gsm if fabric else None),
            'fabric_colour': fabric.colour if fabric else [],
            'fabric_dia': fabric.dia if fabric else [],
        }
        parsed_rows.append(row_dict)
    return filter_stock_rows(parsed_rows, selected_filters or {})


def order_item_process(order, item):
    return {
        'process_type': item.get('process_type') or order.get('process_type') or '',
        'process': item.get('process') or order.get('process') or '',
        'fabric_incharge': item.get('fabric_incharge') or order.get('fabric_incharge') or '',
    }


def generated_order_wip_rows():
    rows = []
    orders = GeneratedOrder.query.filter(
        (GeneratedOrder.status.is_(None)) | (GeneratedOrder.status != 'Completed')
    ).order_by(GeneratedOrder.created_at.desc()).all()
    
    for order in orders:
        order_dict = {
            'id': order.id,
            'po_number': order.po_number,
            'process_type': order.process_type or '',
            'process': order.process or '',
            'fabric_incharge': order.fabric_incharge or '',
            'status': order.status or 'Pending',
        }
        
        for index, item in enumerate(json.loads(order.order_items or '[]')):
            process_data = order_item_process(order_dict, item)
            rows.append({
                'order_id': order.id,
                'item_index': index,
                'po_number': order.po_number,
                'fabric_name': item.get('fabric_name'),
                'uom': item.get('uom'),
                'gsm': item.get('gsm'),
                'colour': item.get('colour', []),
                'dia': item.get('dia', []),
                'quantity': float(item.get('order_qty') or 0),
                'process_type': process_data['process_type'],
                'process': process_data['process'],
                'fabric_incharge': process_data['fabric_incharge'],
                'status': order.status or 'Pending',
            })
    return rows


def get_order_wip_quantity(fabric_name, uom, gsm, colour, dia):
    total = 0.0
    target_colour = [c for c in ([colour] if isinstance(colour, str) else colour or []) if c]
    target_dia = [d for d in ([dia] if isinstance(dia, str) else dia or []) if d]
    for row in generated_order_wip_rows():
        if row.get('fabric_name') != fabric_name:
            continue
        if row.get('uom') != uom or row.get('gsm') != gsm:
            continue
        if target_colour and row.get('colour') != target_colour:
            continue
        if target_dia and row.get('dia') != target_dia:
            continue
        total += row.get('quantity') or 0
    return total


def _stock_entry_matches(entry, colour, dia):
    if colour is not None and colour not in entry.colour:
        return False
    if dia is not None and dia not in entry.dia:
        return False
    return True


def _add_combos(combos, colours, dias, gsm):
    if colours and dias:
        for c in colours:
            for d in dias:
                combos.add((c, d, gsm))
    elif colours:
        for c in colours:
            combos.add((c, None, gsm))
    elif dias:
        for d in dias:
            combos.add((None, d, gsm))
    else:
        combos.add((None, None, gsm))


def _build_requirement_rows_raw():
    fabrics = get_fabrics()
    stock_entries = StockEntry.query.filter(StockEntry.entry_type.in_(['base_stock', 'actual_stock', 'requirement'])).all()
    entries_by_fabric = defaultdict(list)
    for entry in stock_entries:
        entries_by_fabric[entry.fabric_id].append(entry)

    wip_rows = generated_order_wip_rows()
    rows_dict = {}

    for fabric in fabrics:
        fabric_id = fabric['id']
        fabric_name = fabric['name']
        fabric_uom = fabric.get('uom')
        fabric_gsm = fabric.get('gsm')
        combos = set()
        _add_combos(combos, fabric.get('colour', []), fabric.get('dia', []), fabric.get('gsm'))

        for stock_entry in entries_by_fabric.get(fabric_id, []):
            _add_combos(combos, stock_entry.colour, stock_entry.dia, stock_entry.gsm or fabric.get('gsm'))

        
        matching_wip_rows = [
            row for row in wip_rows
            if row.get('fabric_name') == fabric_name
            and row.get('uom') == fabric_uom
        ]

        for wip_row in matching_wip_rows:
            _add_combos(
                combos,
                wip_row.get('colour', []),
                wip_row.get('dia', []),
                wip_row.get('gsm')
            )

        for colour, dia, combo_gsm in combos:
            matching_wip_rows = [
                row for row in wip_rows
                if row.get('fabric_name') == fabric_name
                and row.get('uom') == fabric_uom
                and row.get('gsm') == combo_gsm
                and (
                    (colour is None or row.get('colour') == [colour])
                    and
                    (dia is None or row.get('dia') == [dia])
                )
            ]
            base = 0
            actual = 0
            req_qty = 0
            min_purchase_qty = None

            entry_gsm = fabric.get('gsm')

            for entry in entries_by_fabric.get(fabric_id, []):

                entry_gsm = entry.gsm or fabric.get('gsm')

                if entry_gsm != combo_gsm:
                    continue
                if not _stock_entry_matches(entry, colour, dia):
                    continue
                if entry.entry_type == 'base_stock':
                    base += entry.quantity
                    if entry.min_purchase_qty is not None:
                        min_purchase_qty = entry.min_purchase_qty if min_purchase_qty is None else min(min_purchase_qty, entry.min_purchase_qty)
                elif entry.entry_type == 'actual_stock':
                    actual += entry.quantity
                elif entry.entry_type == 'requirement':
                    req_qty += entry.quantity

            wip = sum(w.get('quantity') or 0 for w in matching_wip_rows)

            key = (
                fabric_name,
                fabric_uom,
                combo_gsm,
                str(colour).strip(),
                str(dia).strip()
            )    

            row = {
                'fabric_id': fabric_id,
                'fabric_name': fabric_name,
                'uom': fabric_uom,
                'gsm': combo_gsm,
                'colour': [colour] if colour is not None else [],
                'dia': [dia] if dia is not None else [],
                'actual_stock': actual,
                'wip': wip,
                'base_stock': base,
                'min_order_qty': min_purchase_qty or 0,
                'requirement_detail': actual + wip - base,
                'quantity': req_qty,
                'note': '',
                'created_at': '',
                'result_type': 'EXCESS' if actual + wip - base > 0 else 'REQUIRED' if actual + wip - base < 0 else 'BALANCED',
            }
            if key not in rows_dict:
                rows_dict[key] = row
            else:
                existing = rows_dict[key]
                existing['actual_stock'] += actual
                existing['wip'] += wip
                existing['base_stock'] += base
                existing['quantity'] += req_qty
                existing['requirement_detail'] = (
                    existing['actual_stock']
                    + existing['wip']
                    - existing['base_stock']
                )
                new_min = row['min_order_qty']
                if new_min and existing['min_order_qty']:
                    existing['min_order_qty'] = min(existing['min_order_qty'], new_min)
                elif new_min:
                    existing['min_order_qty'] = new_min
                existing['result_type'] = (
                    'EXCESS' if existing['requirement_detail'] > 0
                    else 'REQUIRED' if existing['requirement_detail'] < 0
                    else 'BALANCED'
                )

    return list(rows_dict.values())


def build_requirement_rows(selected_filters=None):
    rows = _build_requirement_rows_raw()
    selected_filters = selected_filters or {}
    rows = filter_stock_rows(rows, selected_filters)
    result_filter = filter_values(selected_filters, 'result')
    if result_filter:
        rows = [r for r in rows if matches_filter(r.get('result_type'), result_filter)]
    return rows


def import_stock_excel(file_stream, entry_type, replace=False):
    wb = load_workbook(filename=file_stream, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    mapping = {}
    for index, name in enumerate(headers):
        if 'fabric' in name:
            mapping['fabric'] = index
        elif 'gsm' in name:
            mapping['gsm'] = index  
        elif 'quantity' in name:
            mapping['quantity'] = index
        elif 'note' in name:
            mapping['note'] = index
        elif 'colour' in name:
            mapping['colour'] = index
        elif 'dia' in name:
            mapping['dia'] = index
        elif 'min_purchase' in name or 'min purchase' in name:
            mapping['min_purchase_qty'] = index

    if 'fabric' not in mapping or 'quantity' not in mapping:
        return 'Excel file must include Fabric and Quantity columns.'

    fabric_names = {f.name: f.id for f in Fabric.query.all()}
    new_entries = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[mapping['fabric']]:
            continue

        fabric_name = str(row[mapping['fabric']]).strip()
        quantity_value = row[mapping['quantity']]
        note_value = ''
        colour_value = ''
        dia_value = ''
        min_purchase_qty_value = None

        if 'note' in mapping:
            note_value = str(row[mapping['note']]).strip() if row[mapping['note']] is not None else ''
        if 'colour' in mapping:
            colour_value = str(row[mapping['colour']]).strip() if row[mapping['colour']] is not None else ''
        if 'dia' in mapping:
            dia_value = str(row[mapping['dia']]).strip() if row[mapping['dia']] is not None else ''
        if 'min_purchase_qty' in mapping:
            try:
                min_purchase_qty_value = float(row[mapping['min_purchase_qty']]) if row[mapping['min_purchase_qty']] is not None else None
            except (TypeError, ValueError):
                min_purchase_qty_value = None

        if fabric_name not in fabric_names:
            return f'Fabric not found: {fabric_name}'

        try:
            quantity = float(quantity_value)
        except (TypeError, ValueError):
            return f'Invalid quantity for fabric {fabric_name}. Must be a number.'
        
        gsm_value = str(row[mapping['gsm']]).strip() if 'gsm' in mapping else ''

        entry = StockEntry(
            fabric_id=fabric_names[fabric_name],
            gsm=gsm_value,
            entry_type=entry_type,
            quantity=quantity,
            note=note_value,
            colour=[value.strip() for value in colour_value.split(',') if value.strip()] if colour_value else [],
            dia=[value.strip() for value in dia_value.split(',') if value.strip()] if dia_value else [],
            min_purchase_qty=min_purchase_qty_value,
        )
        new_entries.append(entry)

    if replace:
        StockEntry.query.filter(StockEntry.entry_type == entry_type).delete(synchronize_session=False)

    for entry in new_entries:
        db.session.add(entry)

    db.session.commit()
    return None

def export_stock_excel(entry_type, selected_filters=None):
    wb = Workbook()
    ws = wb.active
    if entry_type == 'requirement':
        ws.append(['Fabric', 'UOM', 'GSM', 'Colour', 'DIA', 'Actual Stock', 'WIP', 'Base Stock', 'Requirement', 'Result'])
        rows = build_requirement_rows(selected_filters)
        for row in rows:
            ws.append([
                row['fabric_name'],
                row['uom'],
                row['gsm'],
                ', '.join(row['colour']) if row['colour'] else '',
                ', '.join(row['dia']) if row['dia'] else '',
                row['actual_stock'],
                row['wip'],
                row['base_stock'],
                row['requirement_detail'],
                row['result_type'],
            ])
    else:
        rows = get_stock_entries(entry_type, selected_filters)
        ws.append(['Fabric', 'UOM', 'GSM', 'Colour', 'DIA', 'Quantity', 'Min Purchase Qty', 'Note', 'Created At', 'Entry Type'])
        for row in rows:
            ws.append([
                row['fabric_name'],
                row['uom'],
                row['gsm'],
                ', '.join(row['colour']) if row['colour'] else '',
                ', '.join(row['dia']) if row['dia'] else '',
                row['quantity'],
                row.get('min_purchase_qty', ''),
                row['note'],
                row['created_at'],
                row['entry_type'],
            ])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream



@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        u_name = (request.form.get("username") or "").strip()
        p = request.form.get("password") or ""
        user = User.query.filter_by(username=u_name).first()
        if user and user.check_password(p):
            session.clear()
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = user.role
            return redirect("/dashboard")
        error = "Invalid username or password"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear(); return redirect("/")


# ---------------- USER MASTER ----------------
@app.route("/users")
@role_required("admin")
def users():
    return render_template("user_master.html",
        users=User.query.order_by(User.id).all(),
        permission_groups=PERMISSION_GROUPS)

@app.route("/users/new", methods=["GET", "POST"])
@role_required("admin")
def user_new():
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        role = (request.form.get("role") or "user").strip().lower()
        if role not in ("admin","user"): role = "user"
        selected = [p for p in request.form.getlist("perms") if p in ALL_PERMISSIONS]
        if not username or not password:
            error = "Username and password are required."
        elif User.query.filter_by(username=username).first():
            error = f"Username '{username}' already exists."
        else:
            u = User(username=username, role=role)
            u.set_password(password)
            u.permissions = ALL_PERMISSIONS if role == "admin" else selected
            db.session.add(u); db.session.commit()
            return redirect("/users")
    return render_template("user_form.html", user=None, error=error,
                           permission_groups=PERMISSION_GROUPS,
                           all_permissions=ALL_PERMISSIONS)

@app.route("/edit_user/<int:user_id>", methods=["GET", "POST"])
@role_required("admin")
def edit_user(user_id):
    u = User.query.get_or_404(user_id); error = None
    if request.method == "POST":
        new_username = (request.form.get("username") or "").strip()
        new_password = request.form.get("password") or ""
        new_role = (request.form.get("role") or u.role).strip().lower()
        selected = [p for p in request.form.getlist("perms") if p in ALL_PERMISSIONS]
        if new_username and new_username != u.username:
            if User.query.filter_by(username=new_username).first():
                error = f"Username '{new_username}' already exists."
            else: u.username = new_username
        if not error and new_role in ("admin","user"):
            me = current_user()
            if me and me.id == u.id and new_role != "admin":
                error = "You can't change your own role from admin."
            else: u.role = new_role
        if not error and new_password: u.set_password(new_password)
        if not error:
            u.permissions = ALL_PERMISSIONS if u.role == "admin" else selected
            db.session.commit(); return redirect("/users")
    return render_template("user_form.html", user=u, error=error,
                           permission_groups=PERMISSION_GROUPS,
                           all_permissions=ALL_PERMISSIONS)

@app.route("/delete_user/<int:user_id>")
@role_required("admin")
def delete_user(user_id):
    me = current_user()
    if me and me.id == user_id: return "You cannot delete your own account.", 400
    u = User.query.get(user_id)
    if u: db.session.delete(u); db.session.commit()
    return redirect("/users")


# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


# ---------------- COLOUR ----------------
@app.route("/colour", methods=["GET", "POST"])
@login_required
def colour():
    u = current_user()
    if not u.has("colour_view"): return "Access denied.", 403
    search = request.args.get("search", "")
    if request.method == "POST":
        if not u.has("colour_add"): return "Access denied.", 403
        c = Colour(name=request.form["colour"], code=request.form["code"])
        db.session.add(c); db.session.commit()
        return redirect("/colour")
    q = Colour.query
    if search: q = q.filter(Colour.name.ilike(f"%{search}%"))
    return render_template("colour_master.html",
        colours=[_to_dict_colour(c) for c in q.order_by(Colour.id).all()])

@app.route("/delete_colour/<int:index>")
@permission_required("colour_delete")
def delete_colour(index):
    c = Colour.query.get(index)
    if c: db.session.delete(c); db.session.commit()
    return redirect("/colour")

@app.route("/edit_colour/<int:index>", methods=["GET", "POST"])
@permission_required("colour_edit")
def edit_colour(index):
    c = Colour.query.get_or_404(index)
    if request.method == "POST":
        c.name = request.form["colour"]; c.code = request.form["code"]
        db.session.commit(); return redirect("/colour")
    return render_template("edit_colour.html", colour=_to_dict_colour(c), index=c.id)


# ---------------- SIZE ----------------
@app.route("/size", methods=["GET", "POST"])
@login_required
def size():
    u = current_user()
    if not u.has("size_view"): return "Access denied.", 403
    search = request.args.get("search", "")
    if request.method == "POST":
        if not u.has("size_add"): return "Access denied.", 403
        s = Size(name=request.form["size"])
        db.session.add(s); db.session.commit()
        return redirect("/size")
    q = Size.query
    if search: q = q.filter(Size.name.ilike(f"%{search}%"))
    return render_template("size_master.html",
        sizes=[_to_dict_size(s) for s in q.order_by(Size.id).all()])

@app.route("/delete_size/<int:index>")
@permission_required("size_delete")
def delete_size(index):
    s = Size.query.get(index)
    if s: db.session.delete(s); db.session.commit()
    return redirect("/size")

@app.route("/edit_size/<int:index>", methods=["GET", "POST"])
@permission_required("size_edit")
def edit_size(index):
    s = Size.query.get_or_404(index)
    if request.method == "POST":
        s.name = request.form["size"]
        db.session.commit(); return redirect("/size")
    return render_template("edit_size.html", size=_to_dict_size(s), index=s.id)


# ---------------- FABRIC ----------------
@app.route("/fabric")
@permission_required("fabric_view")
def fabric():
    return render_template("fabric_master.html",
        fabrics=[_to_dict_fabric(f) for f in Fabric.query.order_by(Fabric.id).all()])

@app.route("/fabric/new", methods=["GET", "POST"])
@permission_required("fabric_add")
def fabric_new():
    if request.method == "POST":
        f = Fabric(name=request.form["name"],
                   uom=request.form.get("uom",""),
                   gsm=request.form.get("gsm",""))
        f.dia = [d.strip() for d in request.form.getlist("dia[]") if d.strip()]
        f.colour = [c.strip() for c in request.form.getlist("colour[]") if c.strip()]
        db.session.add(f); db.session.commit()
        return redirect("/fabric")
    return render_template("fabric_form.html", fabric=None,
        colours=[_to_dict_colour(c) for c in Colour.query.order_by(Colour.id).all()])

@app.route("/edit_fabric/<int:fid>", methods=["GET", "POST"])
@permission_required("fabric_edit")
def edit_fabric(fid):
    f = Fabric.query.get_or_404(fid)
    if request.method == "POST":
        f.name = request.form["name"]
        f.uom = request.form.get("uom","")
        f.gsm = request.form.get("gsm","")
        f.dia = [d.strip() for d in request.form.getlist("dia[]") if d.strip()]
        f.colour = [c.strip() for c in request.form.getlist("colour[]") if c.strip()]
        db.session.commit(); return redirect("/fabric")
    return render_template("fabric_form.html", fabric=_to_dict_fabric(f),
        colours=[_to_dict_colour(c) for c in Colour.query.order_by(Colour.id).all()])

@app.route("/delete_fabric/<int:fid>")
@permission_required("fabric_delete")
def delete_fabric(fid):
    f = Fabric.query.get(fid)
    if f: db.session.delete(f); db.session.commit()
    return redirect("/fabric")


# ---------------- STOCK / ORDER FUNCTIONS ----------------
@app.route('/stock/<entry_type>', methods=['GET', 'POST'])
@login_required
def stock_entries(entry_type):
    u = current_user()
    permission_map = {
        'base_stock': 'base_stock_view',
        'actual_stock': 'actual_stock_view',
        'requirement': 'requirement_view',
    }
    if not u.has(permission_map.get(entry_type, 'program_view')):
        return "Access denied.", 403
    if entry_type not in STOCK_TYPES:
        return redirect('/fabric')

    fabrics = get_fabrics()
    error = None

    if request.method == 'POST':
        if 'upload_excel' in request.form or 'replace_excel' in request.form:
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                error = 'Please choose an Excel file to upload.'
            elif not allowed_excel_file(file.filename):
                error = 'Only .xlsx files are supported.'
            else:
                replace = 'replace_excel' in request.form
                error = import_stock_excel(file, entry_type, replace)
                if not error:
                    return redirect(f'/stock/{entry_type}')
        else:
            fabric_id = request.form.get('fabric_id')
            quantity = request.form.get('quantity', '').strip()
            note = request.form.get('note', '').strip()
            colour_value = request.form.get('colour', '').strip()
            dia_value = request.form.get('dia', '').strip()
            min_purchase_qty = request.form.get('min_purchase_qty', '').strip()
            edit_id = request.form.get('edit_id')

            if not fabric_id or not quantity:
                error = 'Fabric and quantity are required.'
            else:
                try:
                    quantity_value = float(quantity)
                    min_purchase_qty_value = float(min_purchase_qty) if min_purchase_qty else None
                    if edit_id:
                        entry = StockEntry.query.get(edit_id)
                        if entry:
                            entry.fabric_id = int(fabric_id)
                            entry.gsm = request.form.get('gsm', '').strip()
                            entry.quantity = quantity_value
                            entry.note = note
                            entry.colour = [colour_value] if colour_value else []
                            entry.dia = [dia_value] if dia_value else []
                            entry.min_purchase_qty = min_purchase_qty_value
                    else:
                        entry = StockEntry(
                            fabric_id=int(fabric_id),
                            gsm=request.form.get('gsm', '').strip(),
                            entry_type=entry_type,
                            quantity=quantity_value,
                            note=note,
                            colour=[colour_value] if colour_value else [],
                            dia=[dia_value] if dia_value else [],
                            min_purchase_qty=min_purchase_qty_value,
                        )
                        db.session.add(entry)
                    db.session.commit()
                    return redirect(f'/stock/{entry_type}')
                except ValueError:
                    error = 'Quantity must be a number.'

    selected_filters = request_filter_values(['fabric', 'uom', 'gsm', 'colour', 'dia', 'result'])

    if entry_type == 'requirement':
        all_rows = _build_requirement_rows_raw()
        rows = filter_stock_rows(all_rows, selected_filters)
        result_filter = filter_values(selected_filters, 'result')
        if result_filter:
            rows = [r for r in rows if matches_filter(r.get('result_type'), result_filter)]
        total_quantity = sum(r['quantity'] for r in rows)
        filter_options = build_filter_options(all_rows, include_result=True)
    else:
        rows = get_stock_entries(entry_type, selected_filters)
        total_quantity = sum(row['quantity'] for row in rows)
        filter_options = build_filter_options(get_stock_entries(entry_type, {}))

    computed_requirement = None
    actual_stock_total = None
    base_total = None
    wip_total = None
    result_counts = {'REQUIRED': 0, 'EXCESS': 0, 'BALANCED': 0}
    result_totals = {'required_qty': 0.0, 'excess_qty': 0.0}

    if entry_type == 'requirement':
        base_total = db.session.query(db.func.coalesce(db.func.sum(StockEntry.quantity), 0)).filter(
            StockEntry.entry_type == 'base_stock'
        ).scalar() or 0
        wip_total = sum(row.get('quantity') or 0 for row in generated_order_wip_rows())
        actual_stock_total = db.session.query(db.func.coalesce(db.func.sum(StockEntry.quantity), 0)).filter(
            StockEntry.entry_type == 'actual_stock'
        ).scalar() or 0
        computed_requirement = actual_stock_total + wip_total - base_total

        for row in rows:
            row['requirement_detail'] = (row.get('actual_stock') or 0) + (row.get('wip') or 0) - (row.get('base_stock') or 0)
            if row['requirement_detail'] > 0:
                row['result_type'] = 'EXCESS'
            elif row['requirement_detail'] < 0:
                row['result_type'] = 'REQUIRED'
            else:
                row['result_type'] = 'BALANCED'

        for row in rows:
            rt = row.get('result_type')
            if rt in result_counts:
                result_counts[rt] += 1
            if row.get('requirement_detail', 0) > 0:
                result_totals['excess_qty'] += float(row['requirement_detail'])
            elif row.get('requirement_detail', 0) < 0:
                result_totals['required_qty'] += abs(float(row['requirement_detail']))

        base_total = sum((r.get('base_stock') or 0) for r in rows)
        wip_total = sum((r.get('wip') or 0) for r in rows)
        actual_stock_total = sum((r.get('actual_stock') or 0) for r in rows)
        computed_requirement = actual_stock_total + wip_total - base_total

        return render_template(
            'requirement_report.html',
            entry_type=entry_type,
            title=STOCK_TYPES[entry_type],
            entry_label=ENTRY_LABELS[entry_type],
            fabrics=fabrics,
            rows=rows,
            total_quantity=total_quantity,
            actual_stock_total=actual_stock_total,
            base_total=base_total,
            wip_total=wip_total,
            computed_requirement=computed_requirement,
            result_counts=result_counts,
            result_totals=result_totals,
            filter_options=filter_options,
            selected_filters=selected_filters,
            error=error,
        )

    template_name = 'stock_entries.html'
    if entry_type == 'base_stock':
        template_name = 'base_stock_entries.html'
    elif entry_type == 'actual_stock':
        template_name = 'actual_stock_entries.html'
    
    return render_template(
        template_name,
        entry_type=entry_type,
        title=STOCK_TYPES[entry_type],
        entry_label=ENTRY_LABELS[entry_type],
        fabrics=fabrics,
        rows=rows,
        total_quantity=total_quantity,
        actual_stock_total=actual_stock_total,
        base_total=base_total,
        wip_total=wip_total,
        computed_requirement=computed_requirement,
        result_counts=result_counts,
        result_totals=result_totals,
        filter_options=filter_options,
        selected_filters=selected_filters,
        error=error,
    )


@app.route('/stock/<entry_type>/export')
@login_required
def export_stock(entry_type):
    if entry_type not in STOCK_TYPES:
        return redirect('/fabric')
    selected_filters = request_filter_values(['fabric', 'uom', 'gsm', 'colour', 'dia', 'result'])
    file_stream = export_stock_excel(entry_type, selected_filters)
    return send_file(
        file_stream,
        download_name=f'{entry_type}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/stock/<entry_type>/delete', methods=['POST'])
@login_required
def delete_stock_entries(entry_type):
    ids = request.form.getlist('selected_ids')
    if not ids:
        return redirect(f'/stock/{entry_type}')
    clean_ids = []
    for i in ids:
        try:
            clean_ids.append(int(i))
        except (TypeError, ValueError):
            continue
    if not clean_ids:
        return redirect(f'/stock/{entry_type}')
    StockEntry.query.filter(StockEntry.id.in_(clean_ids), StockEntry.entry_type == entry_type).delete(synchronize_session=False)
    db.session.commit()
    return redirect(f'/stock/{entry_type}')


@app.route('/base_stock')
@login_required
def base_stock():
    return redirect('/stock/base_stock')


@app.route('/actual_stock')
@login_required
def actual_stock():
    return redirect('/stock/actual_stock')


@app.route('/requirement')
@login_required
def requirement():
    return redirect('/stock/requirement')


@app.route('/fabric_orders', methods=['GET'])
@login_required
def fabric_orders():
    u = current_user()
    if not u.has('fabric_orders_view'):
        return "Access denied.", 403
    selected_filters = request_filter_values(['fabric', 'uom', 'gsm', 'colour', 'dia', 'result'])
    selected_report_filters = {
        'po_number': request_list_arg('report_po_number'),
        'fabric': request_list_arg('report_fabric'),
        'colour': request_list_arg('report_colour'),
        'dia': request_list_arg('report_dia'),
        'process_type': request_list_arg('report_process_type'),
        'process': request_list_arg('report_process'),
        'fabric_incharge': request_list_arg('report_fabric_incharge'),
        'status': request_list_arg('report_status'),
    }

    filterable_selected_filters = selected_filters
    all_rows = _build_requirement_rows_raw()
    all_results = sorted({r['result_type'] for r in all_rows if r.get('result_type')})
    rows = build_requirement_rows(filterable_selected_filters)
    frozen_orders = session.get('fabric_orders', [])
    frozen_map = {
        (
            f['fabric_name'],
            f['uom'],
            f['gsm'],
            tuple(f['colour']),
            tuple(f['dia'])
        ): float(f.get('order_qty', 0))
        for f in frozen_orders
    }

    for index, row in enumerate(rows):
        row['index'] = index
        key = (
            row['fabric_name'],
            row['uom'],
            row['gsm'],
            tuple(row['colour']),
            tuple(row['dia'])
        )
        row['frozen_qty'] = frozen_map.get(key, 0.0)
        shortage = abs(row['requirement_detail']) if row['requirement_detail'] < 0 else 0.0
        row['order_suggestion'] = max(0.0, shortage - row['min_order_qty'] + row['frozen_qty'])
        row['order_qty'] = row['order_suggestion']

    fabrics = get_fabrics()
    incharges = [{'id': i.id, 'name': i.name} for i in FabricIncharge.query.order_by(FabricIncharge.name).all()]
    processes = [{'id': p.id, 'name': p.name} for p in Process.query.order_by(Process.name).all()]
    process_details = db.session.query(
        ProcessDetails.id,
        ProcessDetails.process_type,
        ProcessDetails.process_id,
        ProcessDetails.fabric_incharge_id,
        Process.name.label('process_name'),
        FabricIncharge.name.label('fabric_incharge_name')
    ).outerjoin(Process).outerjoin(FabricIncharge).order_by(Process.name).all()
    process_details = [
        {
            'id': pd.id,
            'process_type': pd.process_type,
            'process_id': pd.process_id,
            'fabric_incharge_id': pd.fabric_incharge_id,
            'process_name': pd.process_name,
            'fabric_incharge_name': pd.fabric_incharge_name
        }
        for pd in process_details
    ]

    orders = GeneratedOrder.query.order_by(GeneratedOrder.created_at.desc()).all()
    orders_list = []
    for o in orders:
        o_dict = {
            'id': o.id,
            'po_number': o.po_number,
            'process_type': o.process_type,
            'process': o.process,
            'fabric_incharge': o.fabric_incharge,
            'status': o.status,
            'created_at': o.created_at,
            'order_items': json.loads(o.order_items or '[]')
        }
        for item_index, item in enumerate(o_dict['order_items']):
            process_data = order_item_process(o_dict, item)
            item['item_index'] = item_index
            item['process_type'] = process_data['process_type']
            item['process'] = process_data['process']
            item['fabric_incharge'] = process_data['fabric_incharge']
        o_dict['item_count'] = len(o_dict['order_items'])
        o_dict['total_qty'] = sum(float(item.get('order_qty', 0)) for item in o_dict['order_items'])
        orders_list.append(o_dict)

    report_filter_options = {
        'po_number': sorted({o['po_number'] for o in orders_list if o.get('po_number')}),
        'fabric': sorted({item.get('fabric_name') for o in orders_list for item in o['order_items'] if item.get('fabric_name')}),
        'colour': sorted({colour for o in orders_list for item in o['order_items'] for colour in item.get('colour', []) if colour}),
        'dia': sorted({dia for o in orders_list for item in o['order_items'] for dia in item.get('dia', []) if dia}),
        'process_type': sorted({item.get('process_type') for o in orders_list for item in o['order_items'] if item.get('process_type')}),
        'process': sorted({item.get('process') for o in orders_list for item in o['order_items'] if item.get('process')}),
        'fabric_incharge': sorted({item.get('fabric_incharge') for o in orders_list for item in o['order_items'] if item.get('fabric_incharge')}),
        'status': sorted({o['status'] for o in orders_list if o.get('status')}),
    }

    def report_item_matches(item):
        if selected_report_filters['fabric'] and item.get('fabric_name') not in selected_report_filters['fabric']:
            return False
        if selected_report_filters['colour'] and not any(c in selected_report_filters['colour'] for c in item.get('colour', [])):
            return False
        if selected_report_filters['dia'] and not any(d in selected_report_filters['dia'] for d in item.get('dia', [])):
            return False
        if selected_report_filters['process_type'] and item.get('process_type') not in selected_report_filters['process_type']:
            return False
        if selected_report_filters['process'] and item.get('process') not in selected_report_filters['process']:
            return False
        if selected_report_filters['fabric_incharge'] and item.get('fabric_incharge') not in selected_report_filters['fabric_incharge']:
            return False
        return True

    def order_matches_report_filters(order):
        if selected_report_filters['po_number'] and order.get('po_number') not in selected_report_filters['po_number']:
            return False
        if selected_report_filters['status'] and order.get('status') not in selected_report_filters['status']:
            return False

        item_filters_active = any(selected_report_filters[key] for key in ('fabric', 'colour', 'dia', 'process_type', 'process', 'fabric_incharge'))
        return not item_filters_active or any(report_item_matches(item) for item in order['order_items'])

    filtered_orders = []
    item_filters_active = any(selected_report_filters[key] for key in ('fabric', 'colour', 'dia', 'process_type', 'process', 'fabric_incharge'))
    for order in orders_list:
        if not order_matches_report_filters(order):
            continue
        if item_filters_active:
            order = dict(order)
            order['order_items'] = [item for item in order['order_items'] if report_item_matches(item)]
            order['item_count'] = len(order['order_items'])
            order['total_qty'] = sum(float(item.get('order_qty', 0)) for item in order['order_items'])
        filtered_orders.append(order)
    orders_list = filtered_orders

    def rows_excluding(filter_key):
        temp_filters = {k: v for k, v in selected_filters.items() if k != filter_key}
        temp_rows = filter_stock_rows(all_rows, temp_filters)
        result_filter = filter_values(temp_filters, 'result')
        if result_filter:
            temp_rows = [r for r in temp_rows if matches_filter(r.get('result_type'), result_filter)]
        return temp_rows

    all_fabrics = sorted({r['fabric_name'] for r in rows_excluding('fabric')})
    all_uoms = sorted({r['uom'] for r in rows_excluding('uom') if r['uom']})
    all_gsms = sorted({r['gsm'] for r in rows_excluding('gsm') if r['gsm']})
    all_colours = sorted({colour for r in rows_excluding('colour') for colour in r.get('colour', [])})
    all_dias = sorted({dia for r in rows_excluding('dia') for dia in r.get('dia', [])})
    filter_options = {
        'fabric': all_fabrics,
        'uom': all_uoms,
        'gsm': all_gsms,
        'colour': all_colours,
        'dia': all_dias,
        'result': all_results,
    }

    return render_template(
        'fabric_orders.html',
        title='Fabric Orders',
        rows=rows,
        selected_filters=selected_filters,
        all_fabrics=all_fabrics,
        all_uoms=all_uoms,
        all_gsms=all_gsms,
        all_colours=all_colours,
        all_dias=all_dias,
        all_results=all_results,
        filter_options=filter_options,
        frozen_orders=frozen_orders,
        fabrics=fabrics,
        incharges=incharges,
        processes=processes,
        process_details=process_details,
        orders=orders_list,
        report_filter_options=report_filter_options,
        selected_report_filters=selected_report_filters,
        selected_report_form_filters={f'report_{key}': value for key, value in selected_report_filters.items()},
        tab=request.args.get('tab', 'custom'),
        error=request.args.get('error'),
        success=request.args.get('success'),
    )


@app.route('/fabric_orders/freeze', methods=['POST'])
@login_required
def fabric_orders_freeze():
    raw_selected_filters = {
        'fabric': request.form.get('fabric', '').strip(),
        'uom': request.form.get('uom', '').strip(),
        'gsm': request.form.get('gsm', '').strip(),
        'colour': request.form.get('colour', '').strip(),
        'dia': request.form.get('dia', '').strip(),
        'result': request.form.get('result', '').strip(),
    }
    selected_filters = {
        k: [part.strip() for part in v.split(',') if part.strip()]
        for k, v in raw_selected_filters.items()
    }
    rows = build_requirement_rows(selected_filters)
    frozen_orders = session.get('fabric_orders', [])
    
    selected_rows = [int(i) for i in request.form.getlist('selected_rows') if i.isdigit()]
    
    for row_idx, row in enumerate(rows):
        if row_idx in selected_rows:
            qty_str = request.form.get(f"order_qty_{row_idx}", "0")
            try:
                order_qty = float(qty_str)
            except ValueError:
                order_qty = 0.0
                
            frozen_row = {
                'fabric_name': row['fabric_name'],
                'uom': row['uom'],
                'gsm': row['gsm'],
                'colour': row['colour'],
                'dia': row['dia'],
                'order_qty': order_qty,
                'min_order_qty': row['min_order_qty'],
                'result_type': row['result_type'],
            }
            
            exists = False
            for f in frozen_orders:
                if (
                    f['fabric_name'] == frozen_row['fabric_name'] and
                    f['uom'] == frozen_row['uom'] and
                    f['gsm'] == frozen_row['gsm'] and
                    f['colour'] == frozen_row['colour'] and
                    f['dia'] == frozen_row['dia']
                ):
                    f['order_qty'] = frozen_row['order_qty']
                    exists = True
                    break
            if not exists:
                frozen_orders.append(frozen_row)
                
    session['fabric_orders'] = frozen_orders
    return redirect('/fabric_orders?tab=custom&success=Successfully froze %d row(s).' % len(selected_rows))


@app.route('/fabric_orders/generate', methods=['POST'])
@login_required
def fabric_orders_generate():
    po_number = request.form.get('po_number', '').strip()
    process_type = request.form.get('process_type', '').strip()
    process = request.form.get('process', '').strip()
    fabric_incharge = request.form.get('fabric_incharge', '').strip()
    frozen_orders = session.get('fabric_orders', [])
    
    if not po_number or not process_type:
        return redirect('/fabric_orders?tab=custom&error=PO Number and Process Type are required.')
    if not frozen_orders:
        return redirect('/fabric_orders?tab=custom&error=No frozen items to generate an order for.')
    if GeneratedOrder.query.filter_by(po_number=po_number).first():
        return redirect(f'/fabric_orders?tab=custom&error=PO Number {po_number} already exists.')
    try:
        order = GeneratedOrder(
            po_number=po_number,
            order_items=json.dumps(frozen_orders),
            process_type=process_type,
            process=process,
            fabric_incharge=fabric_incharge,
            status='Pending'
        )
        db.session.add(order)
        db.session.commit()
        session['fabric_orders'] = []
        return redirect(f'/fabric_orders?tab=report&success=PO {po_number} generated successfully.')
    except Exception as e:
        return redirect(f'/fabric_orders?tab=custom&error={str(e)}')


@app.route('/fabric_orders/manual/save', methods=['POST'])
@permission_required('fabric_orders_manual_add')
@login_required
def fabric_orders_manual_save():
    action = request.form.get('manual_action', '').strip()
    fabric_ids = request.form.getlist('fabric_id[]')
    colours = request.form.getlist('colour[]')
    dias = request.form.getlist('dia[]')
    quantities = request.form.getlist('quantity[]')
    manual_items = []
    for i in range(len(fabric_ids)):
        fid = fabric_ids[i]; col = colours[i]; dia = dias[i]; qty_str = quantities[i]
        if not fid or not qty_str:
            continue
        try:
            qty = float(qty_str)
        except ValueError:
            continue
        fab = Fabric.query.get(fid)
        if not fab: continue
        manual_items.append({
            'fabric_name': fab.name,
            'uom': fab.uom,
            'gsm': fab.gsm,
            'colour': [col] if col else [],
            'dia': [dia] if dia else [],
            'order_qty': qty,
            'min_order_qty': 0.0,
            'result_type': 'MANUAL',
        })
    if not manual_items:
        return redirect('/fabric_orders?tab=manual&error=Please add at least one item with valid quantity.')
    if action == 'generate_new':
        po_number = request.form.get('po_number', '').strip()
        process_type = request.form.get('process_type', '').strip()
        process = request.form.get('process', '').strip()
        fabric_incharge = request.form.get('fabric_incharge', '').strip()
        if not po_number or not process_type:
            return redirect('/fabric_orders?tab=manual&error=PO Number and Process Type are required.')
        if GeneratedOrder.query.filter_by(po_number=po_number).first():
            return redirect(f'/fabric_orders?tab=manual&error=PO Number {po_number} already exists.')
        try:
            order = GeneratedOrder(
                po_number=po_number,
                order_items=json.dumps(manual_items),
                process_type=process_type,
                process=process,
                fabric_incharge=fabric_incharge,
                status='Pending'
            )
            db.session.add(order)
            db.session.commit()
            return redirect(f'/fabric_orders?tab=report&success=New PO {po_number} generated successfully.')
        except Exception as e:
            return redirect(f'/fabric_orders?tab=manual&error={str(e)}')
    elif action == 'add_to_old':
        po_id = request.form.get('old_po_id', '').strip()
        if not po_id:
            return redirect('/fabric_orders?tab=manual&error=Please select an existing PO.')
        po = GeneratedOrder.query.get(po_id)
        if not po:
            return redirect('/fabric_orders?tab=manual&error=PO not found.')
        old_items = json.loads(po.order_items or '[]')
        for new_item in manual_items:
            found = False
            for old_item in old_items:
                if (
                    old_item['fabric_name'] == new_item['fabric_name'] and
                    old_item['uom'] == new_item['uom'] and
                    old_item['gsm'] == new_item['gsm'] and
                    old_item['colour'] == new_item['colour'] and
                    old_item['dia'] == new_item['dia']
                ):
                    old_item['order_qty'] = float(old_item.get('order_qty', 0)) + new_item['order_qty']
                    found = True
                    break
            if not found:
                old_items.append(new_item)
        po.order_items = json.dumps(old_items)
        db.session.commit()
        return redirect(f'/fabric_orders?tab=report&success=Items successfully appended to PO {po.po_number}.')
    return redirect('/fabric_orders?tab=manual')


@app.route('/fabric_orders/update_status', methods=['POST'])
@login_required
def fabric_orders_update_status():
    status = request.form.get('status', '').strip()
    selected_order_ids = request.form.getlist('selected_order_ids')
    if not selected_order_ids:
        selected_rows = request.form.getlist('selected_rows')
        for row_value in selected_rows:
            parts = row_value.split('|')
            if parts and parts[0].strip():
                selected_order_ids.append(parts[0].strip())
    if not selected_order_ids:
        po_id = request.form.get('id', '').strip()
        if po_id:
            selected_order_ids = [po_id]
    if not selected_order_ids or not status:
        return redirect('/fabric_orders?tab=report&error=Invalid status update parameters.')
    updated = False
    for order_id in set(selected_order_ids):
        po = GeneratedOrder.query.get(order_id)
        if po:
            po.status = status
            updated = True
    if updated:
        db.session.commit()
    return redirect('/fabric_orders?tab=report&success=Status updated successfully.')


@app.route('/fabric_orders/delete_selected', methods=['POST'])
@login_required
def fabric_orders_delete_selected():
    selected_rows = request.form.getlist('selected_rows')
    if not selected_rows:
        return redirect('/fabric_orders?tab=report&error=Please select at least one row to delete.')

    rows_by_order = {}
    for row_value in selected_rows:
        parts = row_value.split('|')
        if len(parts) != 2:
            continue
        order_id_str, item_index_str = parts
        try:
            order_id = int(order_id_str.strip())
            item_index = int(item_index_str.strip())
        except ValueError:
            continue
        rows_by_order.setdefault(order_id, set()).add(item_index)

    if not rows_by_order:
        return redirect('/fabric_orders?tab=report&error=Invalid selected rows.')

    deleted_any = False
    for order_id, item_indices in rows_by_order.items():
        po = GeneratedOrder.query.get(order_id)
        if not po:
            continue
        items = json.loads(po.order_items or '[]')
        remaining_items = [item for idx, item in enumerate(items) if idx not in item_indices]
        if len(remaining_items) != len(items):
            deleted_any = True
            if remaining_items:
                po.order_items = json.dumps(remaining_items)
            else:
                db.session.delete(po)
    if deleted_any:
        db.session.commit()
    return redirect('/fabric_orders?tab=report&success=Selected rows deleted successfully.')


@app.route('/fabric_orders/edit', methods=['POST'])
@login_required
def fabric_orders_edit():
    po_id = request.form.get('id', '').strip()
    po_number = request.form.get('po_number', '').strip()
    process_type = request.form.get('process_type', '').strip()
    process = request.form.get('process', '').strip()
    fabric_incharge = request.form.get('fabric_incharge', '').strip()
    status = request.form.get('status', '').strip()
    item_indices = request.form.getlist('item_index[]')
    quantities = request.form.getlist('item_qty[]')
    if not po_id or not po_number:
        return redirect('/fabric_orders?tab=report&error=PO ID and PO Number are required.')
    po = GeneratedOrder.query.get(po_id)
    if not po:
        return redirect('/fabric_orders?tab=report&error=PO not found.')
    items = json.loads(po.order_items or '[]')
    for idx_str, qty_str in zip(item_indices, quantities):
        try:
            idx = int(idx_str); qty = float(qty_str)
            if 0 <= idx < len(items):
                items[idx]['order_qty'] = qty
        except (ValueError, TypeError):
            continue
    items = [item for item in items if float(item.get('order_qty', 0)) > 0]
    if not items:
        db.session.delete(po); db.session.commit()
        return redirect('/fabric_orders?tab=report&success=PO deleted because all items had 0 quantity.')
    po.po_number = po_number
    po.process_type = process_type
    po.process = process
    po.fabric_incharge = fabric_incharge
    po.status = status
    po.order_items = json.dumps(items)
    db.session.commit()
    return redirect('/fabric_orders?tab=report&success=PO details and quantities updated successfully.')


@app.route('/fabric_orders/delete/<int:order_id>', methods=['POST'])
@login_required
def fabric_orders_delete(order_id):
    po = GeneratedOrder.query.get(order_id)
    if po:
        db.session.delete(po); db.session.commit()
    return redirect('/fabric_orders?tab=report&success=PO deleted successfully.')


@app.route('/fabric_orders/process_update/<int:order_id>/<int:item_index>', methods=['GET', 'POST'])
@login_required
def fabric_orders_process_update(order_id, item_index):
    order = GeneratedOrder.query.get(order_id)
    if not order:
        return redirect('/fabric_orders?tab=report&error=PO not found.')
    items = json.loads(order.order_items or '[]')
    if item_index < 0 or item_index >= len(items):
        return redirect('/fabric_orders?tab=report&error=Order item not found.')
    item = items[item_index]
    current_qty = float(item.get('order_qty') or 0)
    order_dict = {'po_number': order.po_number, 'process_type': order.process_type, 'process': order.process, 'fabric_incharge': order.fabric_incharge}
    process_data = order_item_process(order_dict, item)
    incharges = [{'id': i.id, 'name': i.name} for i in FabricIncharge.query.order_by(FabricIncharge.name).all()]
    processes = [{'id': p.id, 'name': p.name} for p in Process.query.order_by(Process.name).all()]
    process_details = db.session.query(
        ProcessDetails.id,
        ProcessDetails.process_type,
        ProcessDetails.process_id,
        ProcessDetails.fabric_incharge_id,
        Process.name.label('process_name'),
        FabricIncharge.name.label('fabric_incharge_name')
    ).outerjoin(Process).outerjoin(FabricIncharge).order_by(Process.name).all()
    process_details = [{
        'id': pd.id,
        'process_type': pd.process_type,
        'process_id': pd.process_id,
        'fabric_incharge_id': pd.fabric_incharge_id,
        'process_name': pd.process_name,
        'fabric_incharge_name': pd.fabric_incharge_name
    } for pd in process_details]
    if request.method == 'POST':
        move_qty_raw = request.form.get('move_qty', '').strip()
        process_type = request.form.get('process_type', '').strip()
        process = request.form.get('process', '').strip()
        fabric_incharge = request.form.get('fabric_incharge', '').strip()
        try:
            move_qty = float(move_qty_raw)
        except ValueError:
            move_qty = 0
        if move_qty <= 0 or move_qty > current_qty:
            return redirect(f'/fabric_orders/process_update/{order_id}/{item_index}?error=Split quantity must be greater than 0 and not more than current quantity.')
        if not process_type:
            return redirect(f'/fabric_orders/process_update/{order_id}/{item_index}?error=Process Type is required.')
        if move_qty == current_qty:
            item['process_type'] = process_type
            item['process'] = process
            item['fabric_incharge'] = fabric_incharge
        else:
            item['order_qty'] = current_qty - move_qty
            moved_item = dict(item)
            moved_item['order_qty'] = move_qty
            moved_item['process_type'] = process_type
            moved_item['process'] = process
            moved_item['fabric_incharge'] = fabric_incharge
            moved_item['split_from'] = item_index
            items.append(moved_item)
        order.order_items = json.dumps(items)
        db.session.commit()
        return redirect('/fabric_orders?tab=report&success=Process movement updated successfully.')
    item = dict(item)
    item.update(process_data)
    return render_template(
        'fabric_order_process_update.html',
        title='Process Update',
        order={'id': order.id, 'po_number': order.po_number},
        item=item,
        item_index=item_index,
        current_qty=current_qty,
        process_details=process_details,
        processes=processes,
        incharges=incharges,
        error=request.args.get('error'),
    )


@app.route('/fabric_orders_report')
@login_required
def fabric_orders_report():
    return redirect('/fabric_orders?tab=report')


@app.route('/fabric_orders/report/export')
@login_required
def fabric_orders_report_export():
    selected_report_filters = {
        'po_number': request_list_arg('report_po_number'),
        'fabric': request_list_arg('report_fabric'),
        'colour': request_list_arg('report_colour'),
        'dia': request_list_arg('report_dia'),
        'process_type': request_list_arg('report_process_type'),
        'process': request_list_arg('report_process'),
        'fabric_incharge': request_list_arg('report_fabric_incharge'),
        'status': request_list_arg('report_status'),
    }
    orders = GeneratedOrder.query.order_by(GeneratedOrder.created_at.desc()).all()
    export_rows = []
    for order in orders:
        order_dict = {'po_number': order.po_number, 'process_type': order.process_type, 'process': order.process, 'fabric_incharge': order.fabric_incharge, 'status': order.status}
        items = json.loads(order.order_items or '[]')
        if selected_report_filters['po_number'] and order.po_number not in selected_report_filters['po_number']:
            continue
        if selected_report_filters['process_type'] and order.process_type not in selected_report_filters['process_type']:
            continue
        if selected_report_filters['process'] and order.process not in selected_report_filters['process']:
            continue
        if selected_report_filters['fabric_incharge'] and order.fabric_incharge not in selected_report_filters['fabric_incharge']:
            continue
        if selected_report_filters['status'] and order.status not in selected_report_filters['status']:
            continue
        for item in items:
            item.update(order_item_process(order_dict, item))
            if selected_report_filters['fabric'] and item.get('fabric_name') not in selected_report_filters['fabric']:
                continue
            if selected_report_filters['colour'] and not any(c in selected_report_filters['colour'] for c in item.get('colour', [])):
                continue
            if selected_report_filters['dia'] and not any(d in selected_report_filters['dia'] for d in item.get('dia', [])):
                continue
            if selected_report_filters['process_type'] and item.get('process_type') not in selected_report_filters['process_type']:
                continue
            if selected_report_filters['process'] and item.get('process') not in selected_report_filters['process']:
                continue
            if selected_report_filters['fabric_incharge'] and item.get('fabric_incharge') not in selected_report_filters['fabric_incharge']:
                continue
            export_rows.append([
                order.po_number or '',
                item.get('fabric_name', ''),
                item.get('uom', ''),
                item.get('gsm', ''),
                ', '.join(item.get('colour', [])),
                ', '.join(item.get('dia', [])),
                item.get('order_qty', ''),
                item.get('process_type', ''),
                item.get('process', ''),
                item.get('fabric_incharge', ''),
                order.status or '',
                order.created_at.isoformat() if order.created_at else '',
            ])
    wb = Workbook(); ws = wb.active; ws.title = 'Fabric Orders'
    ws.append(['PO No','Fabric','UOM','GSM','Colour','DIA','Order Qty','Process Type','Process','Fabric Incharge','Status','Created At'])
    for row in export_rows: ws.append(row)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 32)
    file_stream = BytesIO(); wb.save(file_stream); file_stream.seek(0)
    return send_file(file_stream, as_attachment=True, download_name='fabric_orders_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/process_master')
@login_required
def process_master():
    u = current_user()
    if not u.has('process_view'): return 'Access denied.', 403
    error = request.args.get('error')
    success = request.args.get('success')
    incharges = [{'id': i.id, 'name': i.name} for i in FabricIncharge.query.order_by(FabricIncharge.name).all()]
    processes = [{'id': p.id, 'name': p.name} for p in Process.query.order_by(Process.name).all()]
    details = db.session.query(ProcessDetails.id, ProcessDetails.process_type, ProcessDetails.process_id, ProcessDetails.fabric_incharge_id, Process.name.label('process_name'), FabricIncharge.name.label('fabric_incharge_name')).outerjoin(Process).outerjoin(FabricIncharge).order_by(ProcessDetails.id.desc()).all()
    details = [{
        'id': d.id,
        'process_type': d.process_type,
        'process_id': d.process_id,
        'fabric_incharge_id': d.fabric_incharge_id,
        'process_name': d.process_name,
        'fabric_incharge_name': d.fabric_incharge_name
    } for d in details]
    return render_template('process_master.html', title='Process Master', incharges=incharges, processes=processes, details=details, error=error, success=success)


@app.route('/process_master/fabric_incharge/save', methods=['POST'])
@login_required
def fabric_incharge_save():
    name = request.form.get('name', '').strip()
    edit_id = request.form.get('id', '').strip()
    if not name:
        return redirect('/process_master?error=Name is required')
    if edit_id:
        incharge = FabricIncharge.query.get(edit_id)
        if incharge:
            incharge.name = name
            msg = 'Fabric Incharge updated successfully'
        else:
            return redirect('/process_master?error=Fabric Incharge not found')
    else:
        incharge = FabricIncharge(name=name)
        db.session.add(incharge)
        msg = 'Fabric Incharge added successfully'
    try:
        db.session.commit()
        return redirect(f'/process_master?success={msg}')
    except Exception as e:
        return redirect('/process_master?error=Fabric Incharge name already exists')


@app.route('/process_master/fabric_incharge/delete/<int:incharge_id>', methods=['POST'])
@login_required
def fabric_incharge_delete(incharge_id):
    incharge = FabricIncharge.query.get(incharge_id)
    if incharge:
        db.session.delete(incharge)
        db.session.commit()
    return redirect('/process_master?success=Fabric Incharge deleted successfully')


@app.route('/process_master/process/save', methods=['POST'])
@login_required
def process_save():
    name = request.form.get('name', '').strip()
    edit_id = request.form.get('id', '').strip()
    if not name:
        return redirect('/process_master?error=Process Name is required')
    if edit_id:
        process = Process.query.get(edit_id)
        if process:
            process.name = name
            msg = 'Process updated successfully'
        else:
            return redirect('/process_master?error=Process not found')
    else:
        process = Process(name=name)
        db.session.add(process)
        msg = 'Process added successfully'
    try:
        db.session.commit()
        return redirect(f'/process_master?success={msg}')
    except Exception as e:
        return redirect('/process_master?error=Process name already exists')


@app.route('/process_master/process/delete/<int:process_id>', methods=['POST'])
@login_required
def process_delete(process_id):
    process = Process.query.get(process_id)
    if process:
        db.session.delete(process)
        db.session.commit()
    return redirect('/process_master?success=Process deleted successfully')


@app.route('/process_master/process_details/save', methods=['POST'])
@login_required
def process_details_save():
    process_type = request.form.get('process_type', '').strip()
    process_id = request.form.get('process_id', '').strip()
    fabric_incharge_id = request.form.get('fabric_incharge_id', '').strip()
    edit_id = request.form.get('id', '').strip()
    if not process_type or not process_id or not fabric_incharge_id:
        return redirect('/process_master?error=All fields are required')
    if edit_id:
        detail = ProcessDetails.query.get(edit_id)
        if detail:
            detail.process_type = process_type
            detail.process_id = process_id
            detail.fabric_incharge_id = fabric_incharge_id
            msg = 'Process Details updated successfully'
        else:
            return redirect('/process_master?error=Process Details not found')
    else:
        existing = ProcessDetails.query.filter_by(process_type=process_type, process_id=process_id, fabric_incharge_id=fabric_incharge_id).first()
        if existing:
            return redirect('/process_master?error=This process combination already exists')
        detail = ProcessDetails(process_type=process_type, process_id=process_id, fabric_incharge_id=fabric_incharge_id)
        db.session.add(detail)
        msg = 'Process Details added successfully'
    db.session.commit()
    return redirect(f'/process_master?success={msg}')


@app.route('/process_master/process_details/delete/<int:details_id>', methods=['POST'])
@login_required
def process_details_delete(details_id):
    detail = ProcessDetails.query.get(details_id)
    if detail:
        db.session.delete(detail)
        db.session.commit()
    return redirect('/process_master?success=Process Details deleted successfully')


# ---------------- PRODUCT ----------------
@app.route("/product")
@permission_required("product_view")
def product():
    search = request.args.get("search", "")
    q = Product.query
    if search: q = q.filter(Product.name.ilike(f"%{search}%"))
    return render_template("product_master.html",
        products=[_to_dict_product(p) for p in q.order_by(Product.id).all()])

@app.route("/product/new", methods=["GET", "POST"])
@permission_required("product_add")
def product_new():
    if request.method == "POST":
        p = Product(name=request.form["name"],
                    brand=request.form.get("brand",""),
                    category=request.form.get("category",""),
                    type=request.form.get("type",""),
                    fabric=request.form.get("fabric",""))
        p.colors = [c for c in request.form.getlist("colors[]") if c]
        p.sizes = [s for s in request.form.getlist("sizes[]") if s]
        db.session.add(p); db.session.commit()
        return redirect("/product")
    return render_template("product_form.html", product=None,
        fabrics=[_to_dict_fabric(f) for f in Fabric.query.order_by(Fabric.id).all()],
        colours=[_to_dict_colour(c) for c in Colour.query.order_by(Colour.id).all()],
        sizes=[_to_dict_size(s) for s in Size.query.order_by(Size.id).all()])

@app.route("/edit_product/<int:pid>", methods=["GET", "POST"])
@permission_required("product_edit")
def edit_product(pid):
    p = Product.query.get_or_404(pid)
    if request.method == "POST":
        p.name = request.form["name"]
        p.brand = request.form.get("brand","")
        p.category = request.form.get("category","")
        p.type = request.form.get("type","")
        p.fabric = request.form.get("fabric","")
        p.colors = [c for c in request.form.getlist("colors[]") if c]
        p.sizes = [s for s in request.form.getlist("sizes[]") if s]
        db.session.commit(); return redirect("/product")
    return render_template("product_form.html", product=_to_dict_product(p),
        fabrics=[_to_dict_fabric(f) for f in Fabric.query.order_by(Fabric.id).all()],
        colours=[_to_dict_colour(c) for c in Colour.query.order_by(Colour.id).all()],
        sizes=[_to_dict_size(s) for s in Size.query.order_by(Size.id).all()])

@app.route("/delete_product/<int:pid>")
@permission_required("product_delete")
def delete_product(pid):
    p = Product.query.get(pid)
    if p: db.session.delete(p); db.session.commit()
    return redirect("/product")

@app.route("/get-colors/<path:fabric>")
@login_required
def get_colors(fabric):
    f_name = fabric.split(" (")[0]
    f = Fabric.query.filter_by(name=f_name).first()
    if not f: return jsonify([])
    return jsonify(f.colour)


# ---------------- SUPPLIER MASTER ----------------
@app.route("/supplier", methods=["GET", "POST"])
@login_required
def supplier():
    u = current_user()
    if not u.has("supplier_view"): return "Access denied.", 403
    if request.method == "POST":
        if not u.has("supplier_add"): return "Access denied.", 403
        s = Supplier(
            name=(request.form.get("name") or "").strip(),
            code=(request.form.get("code") or "").strip(),
            contact=(request.form.get("contact") or "").strip(),
            address=(request.form.get("address") or "").strip(),
        )
        if s.name:
            db.session.add(s); db.session.commit()
        return redirect("/supplier")

    search = request.args.get("search", "").strip()
    contact = request.args.get("contact", "").strip()
    code = request.args.get("code", "").strip()
    q = Supplier.query
    if search: q = q.filter(Supplier.name.ilike(f"%{search}%"))
    if contact: q = q.filter(Supplier.contact.ilike(f"%{contact}%"))
    if code:    q = q.filter(Supplier.code.ilike(f"%{code}%"))
    suppliers = [_to_dict_supplier(s) for s in q.order_by(Supplier.id).all()]
    return render_template("supplier_master.html", suppliers=suppliers,
                           f_search=search, f_contact=contact, f_code=code)

@app.route("/edit_supplier/<int:sid>", methods=["GET", "POST"])
@permission_required("supplier_edit")
def edit_supplier(sid):
    s = Supplier.query.get_or_404(sid)
    if request.method == "POST":
        s.name = (request.form.get("name") or "").strip()
        s.code = (request.form.get("code") or "").strip()
        s.contact = (request.form.get("contact") or "").strip()
        s.address = (request.form.get("address") or "").strip()
        db.session.commit(); return redirect("/supplier")
    return render_template("supplier_form.html", supplier=_to_dict_supplier(s))

@app.route("/delete_supplier/<int:sid>")
@permission_required("supplier_delete")
def delete_supplier(sid):
    s = Supplier.query.get(sid)
    if s: db.session.delete(s); db.session.commit()
    return redirect("/supplier")


# ---------------- PROGRAM ----------------
@app.route("/program", methods=["GET", "POST"])
@login_required
def program():

    u = current_user()

    if not u.has("program_view"):
        return "Access denied.", 403

    if request.method == "POST":

        if not u.has("program_add"):
            return "Access denied.", 403

        try:

            program_no = _next_program_no()

            date_str = datetime.now().strftime("%d-%m-%Y")

            fabric = request.form.get("fabric", "")
            dia = request.form.get("dia", "")
            ptype = request.form.get("type", "")
            product = request.form.get("product", "")

            sizes = request.form.getlist("sizes[]")
            ratios = request.form.getlist("ratios[]")

            colors = request.form.getlist("colors[]")
            rolls = request.form.getlist("rolls[]")

            for ci, color in enumerate(colors):

                roll_val = rolls[ci] if ci < len(rolls) else ""

                for si, sz in enumerate(sizes):

                    ratio_val = ratios[si] if si < len(ratios) else ""

                    row = Program(
                        id=str(uuid.uuid4()),
                        program_no=program_no,
                        date=date_str,
                        fabric=fabric,
                        dia=dia,
                        type=ptype,
                        product=product,
                        color=color,
                        size=sz,
                        ratio=ratio_val,
                        rolls=roll_val,
                        status="pending",
                    )

                    db.session.add(row)

            db.session.commit()

            return redirect("/program")

        except Exception as e:

            db.session.rollback()

            print("PROGRAM SAVE ERROR:", e)

            return f"Error saving program: {str(e)}"

    return render_template(
        "program_entry.html",
        fabrics=[_to_dict_fabric(f) for f in Fabric.query.order_by(Fabric.id).all()],
        products=[_to_dict_product(p) for p in Product.query.order_by(Product.id).all()],
        programs=_grouped_programs()
    )



@app.route("/program/<program_no>")
@login_required
def program_view(program_no):
    rows = Program.query.filter_by(program_no=program_no).all()
    if not rows: return "Program not found", 404
    header = {
        "program_no": rows[0].program_no, "date": rows[0].date,
        "fabric": rows[0].fabric, "dia": rows[0].dia,
        "product": rows[0].product,
        "supplier": rows[0].supplier or "",
        "lot_no": rows[0].lot_no or "",
        "status": (rows[0].status or "pending").lower(),
        "completed_date": rows[0].completed_date or "",
    }
    size_ratio = {}; colour_rolls = {}
    for r in rows:
        size_ratio.setdefault(r.size, r.ratio)
        colour_rolls.setdefault(r.color, r.rolls)
    mode = request.args.get("mode", "")
    suppliers = []
    if mode == "allocate":
        suppliers = [_to_dict_supplier(s) for s in Supplier.query.order_by(Supplier.name).all()]
    return render_template("program_view.html",
        header=header, size_ratio=size_ratio, colour_rolls=colour_rolls,
        mode=mode, suppliers=suppliers)


@app.route("/allocate_supplier", methods=["POST"])
@permission_required("program_status")
def allocate_supplier():
    program_no = request.form.get("program_no", "").strip()
    supplier_name = request.form.get("supplier", "").strip()
    if not program_no or not supplier_name:
        return "Program No and Supplier are required.", 400
    rows = Program.query.filter_by(program_no=program_no).all()
    if not rows: return "Program not found.", 404
    for r in rows:
        r.supplier = supplier_name
        r.status = "wip"
    db.session.commit()
    return redirect(f"/program/{program_no}/print")


@app.route("/program/<program_no>/print")
@login_required
def program_print(program_no):
    rows = Program.query.filter_by(program_no=program_no).all()
    if not rows: return "Program not found", 404
    header = {
        "program_no": rows[0].program_no, "date": rows[0].date,
        "fabric": rows[0].fabric, "dia": rows[0].dia,
        "product": rows[0].product,
        "supplier": rows[0].supplier or "",
        "status": (rows[0].status or "pending").lower(),
    }
    size_ratio = {}; colour_rolls = {}
    for r in rows:
        size_ratio.setdefault(r.size, r.ratio)
        colour_rolls.setdefault(r.color, r.rolls)
    return render_template("program_print.html",
        header=header, size_ratio=size_ratio, colour_rolls=colour_rolls)


@app.route("/cutting_completion", methods=["POST"])
@permission_required("program_status")
def cutting_completion():
    program_no = request.form.get("program_no", "").strip()
    lot_no = request.form.get("lot_no", "").strip()
    if not program_no or not lot_no:
        return "Program No and Lot No are required.", 400
    rows = Program.query.filter_by(program_no=program_no).all()
    if not rows: return "Program not found.", 404
    today = datetime.now().strftime("%d-%m-%Y")
    for r in rows:
        r.lot_no = lot_no
        r.status = "completed"
        r.completed_date = today
    db.session.commit()
    return redirect(request.referrer or "/overall_programs")


@app.route("/edit_program/<pid>", methods=["GET", "POST"])
@permission_required("program_edit")
def edit_program(pid):
    p = Program.query.get_or_404(pid)
    if request.method == "POST":
        p.ratio = request.form.get("ratio", p.ratio)
        p.rolls = request.form.get("rolls", p.rolls)
        new_status = (request.form.get("status") or p.status or "pending").lower()
        if new_status not in ("pending","wip","completed"):
            new_status = "pending"
        p.status = new_status
        db.session.commit()
        return redirect("/program")
    return render_template("edit_program.html", p=p)


@app.route("/edit_program_group", methods=["POST"])
@login_required
def edit_program_group():
    u = current_user()
    if not (u.has("program_status") or u.has("overall_status")):
        return "Access denied.", 403
    ids_csv = request.form.get("ids", "")
    new_status = (request.form.get("status") or "pending").lower()
    if new_status not in ("pending","wip","completed"):
        new_status = "pending"
    ids = [i for i in ids_csv.split(",") if i]
    if ids:
        Program.query.filter(Program.id.in_(ids)).update(
            {"status": new_status}, synchronize_session=False)
        db.session.commit()
    return redirect(request.referrer or "/overall_programs")


@app.route("/delete_program_group")
@permission_required("program_delete")
def delete_program_group():
    ids_csv = request.args.get("ids", "")
    ids = [i for i in ids_csv.split(",") if i]
    if ids:
        Program.query.filter(Program.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
    return redirect(request.referrer or "/program")


# ---------------- OVERALL PROGRAMS ----------------
@app.route("/overall_programs")
@permission_required("overall_view")
def overall_programs():
    return render_template("overall_programs.html", programs=_grouped_programs())


# ---------------- STARTUP MIGRATION + ADMIN SEED ----------------
def _safe_migrate():
    with db.engine.begin() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS permissions_csv TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE programs ADD COLUMN IF NOT EXISTS supplier VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE programs ADD COLUMN IF NOT EXISTS lot_no VARCHAR(50) DEFAULT ''"))
        conn.execute(text("ALTER TABLE programs ADD COLUMN IF NOT EXISTS completed_date VARCHAR(20) DEFAULT ''"))
        conn.execute(text("CREATE SEQUENCE IF NOT EXISTS program_no_seq START 1"))
        conn.execute(text(
            "SELECT setval('program_no_seq', COALESCE(MAX(NULLIF(regexp_replace(program_no, '\\D', '', 'g'), '')::integer), 1), true)"
            " FROM programs"
        ))

def _seed_admin():
    if not User.query.filter_by(username="admin").first():
        admin = User(username="admin", role="admin")
        admin.set_password("admin123")
        admin.permissions = ALL_PERMISSIONS
        db.session.add(admin); db.session.commit()

with app.app_context():
    db.create_all()
    try: _safe_migrate()
    except Exception as e: print("Migration warning:", e)
    _seed_admin()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)